# -*- coding: utf-8 -*-
"""构建《项目度量与需求分析.xlsx》分析工作簿。
数据源: data/analysis.json（源自 srs_full.txt / metrics1_dump.txt / metrics2_dump.txt）
用法: python3 build_xlsx.py [输出路径] [工作量偏差阈值覆盖值]
"""
import json, math, os, sys

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import FormulaRule, CellIsRule

BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DATA = os.path.join(BASE, 'data', 'analysis.json')
OUT = sys.argv[1] if len(sys.argv) > 1 else os.path.join(BASE, '项目度量与需求分析.xlsx')
EFFORT_THR = float(sys.argv[2]) if len(sys.argv) > 2 else 0.15

with open(DATA, encoding='utf-8') as f:
    D = json.load(f)
MET = D['metrics']; P1 = MET['projects'][0]; P2 = MET['projects'][1]
DEC = D['decomposition']; QUA = D['quality']; MON = MET['monitoring']

AR = 'Arial'; SZ = 10
F_TX = Font(name=AR, size=SZ)                                # 说明文字
F_IN = Font(name=AR, size=SZ, color='0000FF')                # 蓝=输入
F_FM = Font(name=AR, size=SZ, color='000000')                # 黑=公式
F_XR = Font(name=AR, size=SZ, color='008000')                # 绿=跨表引用
F_HD = Font(name=AR, size=SZ, bold=True, color='FFFFFF')     # 表头
F_TI = Font(name=AR, size=12, bold=True, color='1F4E79')     # 大标题
F_BT = Font(name=AR, size=SZ, bold=True, color='1F4E79')     # 块标题
F_B  = Font(name=AR, size=SZ, bold=True)
F_FMB = Font(name=AR, size=SZ, bold=True, color='000000')
F_INB = Font(name=AR, size=SZ, bold=True, color='0000FF')
F_NT = Font(name=AR, size=9, italic=True, color='808080')    # 灰注
FILL_HD = PatternFill('solid', start_color='1F4E79')
FILL_TOT = PatternFill('solid', start_color='DDEBF7')
FILL_IN = None
WRAP = Alignment(wrap_text=True, vertical='top')
CTR = Alignment(horizontal='center', vertical='center', wrap_text=True)
LFT = Alignment(horizontal='left', vertical='top', wrap_text=True)
thin = Side(style='thin', color='B0B0B0')
BD = Border(left=thin, right=thin, top=thin, bottom=thin)
PCT = '0.0%'; MONEY = '#,##0.00'; N2 = '0.00'; N3 = '0.000'; INT = '0'; N1 = '0.0'

def put(ws, r, c, v, font=F_TX, fmt=None, align=None, fill=None, bd=False):
    cell = ws.cell(row=r, column=c, value=v); cell.font = font
    if fmt: cell.number_format = fmt
    if align: cell.alignment = align
    if fill: cell.fill = fill
    if bd: cell.border = BD
    return cell

def header_row(ws, r, cols, start=1):
    for i, h in enumerate(cols):
        put(ws, r, start + i, h, F_HD, align=CTR, fill=FILL_HD, bd=True)

def widths(ws, wd):
    for col, w in wd.items():
        ws.column_dimensions[col].width = w

def est_h(texts, cpls, min_h=16, max_h=400):
    lines = 1
    for t, cpl in zip(texts, cpls):
        if not t: continue
        t = str(t)
        n = sum(max(1, math.ceil(len(seg) / cpl)) for seg in t.split('\n'))
        lines = max(lines, n)
    return min(max_h, max(min_h, lines * 13.6))

wb = Workbook()

# ============================== 00 说明 ==============================
ws = wb.active; ws.title = '00说明'
widths(ws, {'A': 3, 'B': 110})
ws.freeze_panes = 'A3'
r = 1
put(ws, r, 2, '《供水营销管理系统》项目度量与需求分析工作簿 — 使用说明', F_TI); r += 2
sec = lambda t: put(ws, r, 2, t, F_BT)
def line(t, font=F_TX, cpl=52):
    global r
    put(ws, r, 2, t, font, align=WRAP)
    ws.row_dimensions[r].height = est_h([t], [cpl]); r += 1
sec('一、工作簿用途'); r += 1
line('本工作簿汇集《供水营销管理系统》需求规格说明书（SRS）的功能分解与质量检查结果，以及两个已结项软件项目的度量数据，'
     '用于：①跨项目度量对比与组织级基线参考；②按可调阈值的红绿灯偏差监控；③新项目（24个二级功能）估算与需求评审支撑。')
r += 1; sec('二、工作表清单'); r += 1
for s, t in [
    ('01参数', '监控阈值与口径参数（全部蓝字可修改），缺陷加权权重、人时换算也在此维护，供全簿公式引用'),
    ('02项目1度量', '项目1（度量表1，256FP，2025-02-24~06-30）：规模/工作量/进度/成本四块 + 关键派生指标（供04/05引用）'),
    ('03项目2度量', '项目2（度量表2，KSSH-1102，327FP，2024-09-13~2025-04-02）：工作量/规模效率/工期/成本/缺陷/评审 + 关键派生指标'),
    ('04汇总对比', '两项目15项指标横向对比，全部跨表引用公式取数，附差异说明'),
    ('05监控报告', '红绿灯仪表盘：11项指标 × 01参数阈值自动判定（●正常/▲预警/✖超限），底部自动生成结论与风险建议'),
    ('06功能分解', 'SRS功能分解WBS全清单（8模块27功能，含3个银行代收接口），含优先级映射、输入输出、验收要点；下方附范围缺口扩展建议'),
    ('07质量检查', 'SRS质量问题清单41项（高18/中16/低7），按严重度与类别自动统计，支持筛选'),
]:
    line('「%s」 %s' % (s, t))
r += 1; sec('三、使用方法'); r += 1
for t in [
    '颜色约定：蓝字=可修改的输入原值；黑字=公式自动计算（请勿覆盖）；绿字=跨表引用公式；深蓝底白字=表头。',
    '调整监控口径：修改「01参数」中任一蓝字阈值（如把工作量偏差控制范围由15%改为10%），保存后「05监控报告」的状态列、'
    '统计与自动结论文字即联动重判；「02/03」表中显示的控制范围/目标值文字亦随之更新。',
    '本工作簿公式为标准Excel函数（SUM/IF/COUNTIF/SUMPRODUCT/TEXTJOIN等），在 Excel / WPS / LibreOffice 中打开会自动重算。',
    '「06功能分解」「07质量检查」首行已冻结并开启自动筛选，可按模块、优先级、严重度、类别过滤。',
]:
    line(t)
r += 1; sec('四、关键口径说明'); r += 1
for t in [
    '项目1工作量偏差率以「需求阶段」估计值为基准：偏差率=(实际-需求阶段估计)/需求阶段估计。',
    '项目2缺陷加权数权重经原始数据反推核对为：致命×8、严重×4、一般×1、细微×0.5（102/28/130三行全部吻合），'
    '与常见的10/5/2/1模板口径不同；权重已参数化到「01参数」，如组织口径调整可直接修改。',
    '缺陷探测率=集成与系统测试缺陷加权数/总缺陷加权数；缺陷密度=缺陷加权数/实际规模FP；评审效率=评审缺陷总数/评审用时(人时)。',
    '功能点耗时率=总实际工作量(人天)×8(人时/人天)/实际规模FP；项目1原表未给出该指标，为同口径推算值。',
    '项目2工期为工作日口径（已扣节假日，日历跨度计划189天/实际202天）；项目1总工期为阶段天数合计125天（日历跨度127天）。',
]:
    line(t)
r += 1; sec('五、数据来源与生成信息'); r += 1
for t in [
    '需求规格说明书：水务系统需求规格说明书（解析文本 data/srs_full.txt，7个一级模块、24个二级功能、3个银行代收接口）。',
    '项目1度量数据：项目度量数据表-1（解析文本 data/metrics1_dump.txt，项目属性/规模/工作量/进度/成本5表）。',
    '项目2度量数据：02项目度量数据表（解析文本 data/metrics2_dump.txt，编号KSSH-1102 V1.5受控，工作量规模效率/工期/成本/缺陷4表）。',
    '分析中间结果：data/analysis.json。生成日期：2026-07-09；编制：供水集团科技开发部（需求分析演示项目）。',
]:
    line(t)
r += 1; sec('六、数据质量说明（引用原始数据时请注意）'); r += 1
for i, t in enumerate(MET['dataQualityNotes'], 1):
    line('%d. %s' % (i, t))

# ============================== 01 参数 ==============================
ws = wb.create_sheet('01参数')
widths(ws, {'A': 34, 'B': 10, 'C': 12, 'D': 20, 'E': 12, 'F': 66})
ws.freeze_panes = 'A4'
put(ws, 1, 1, '监控阈值参数（本表蓝字均可修改，修改后 02/03/05 各表自动联动）', F_TI)
put(ws, 2, 1, '蓝字=输入；其余表通过 =\'01参数\'!$C$n 引用本表。', F_NT)
header_row(ws, 3, ['参数名称', '符号', '数值', '判定方向/口径', '默认值', '说明'])
params = [
    ('工作量偏差控制范围', 'thr_effort', EFFORT_THR, '|偏差率|≤阈值', '±15%', '实际总工作量相对估计值的允许偏差', PCT),
    ('工期(进度)偏差控制范围', 'thr_schedule', 0.15, '|偏差率|≤阈值', '±15%', '进度/累计工期偏差率允许范围', PCT),
    ('成本偏差控制范围', 'thr_cost', 0.15, '|偏差率|≤阈值', '±15%', '成本偏差率允许范围', PCT),
    ('规模偏差控制范围', 'thr_size', 0.10, '|偏差率|≤阈值', '±10%', '功能点规模偏差率允许范围', PCT),
    ('缺陷探测率目标', 'thr_detect', 0.75, '指标≥目标', '≥75%', '系统测试加权缺陷占总加权缺陷比例的下限', PCT),
    ('功能点耗时率目标(人时/FP)', 'thr_hourfp', 7.15, '指标≤目标', '≤7.15', '每功能点耗用人时上限', N2),
    ('评审效率目标(个/人时)', 'thr_review', 0.75, '指标≥目标', '≥0.75', '评审每人时发现缺陷数下限', N2),
    ('红绿灯规则系数', 'coef_green', 0.6, '绿区=阈值×系数', '0.6', '|偏差|≤阈值×系数为●正常；≤阈值为▲预警；>阈值为✖超限', N2),
]
for i, (name, sym, val, d, dv, note, fmt) in enumerate(params):
    r = 4 + i
    put(ws, r, 1, name, F_TX, bd=True); put(ws, r, 2, sym, F_NT, bd=True)
    put(ws, r, 3, val, F_IN, fmt=fmt, align=CTR, bd=True)
    put(ws, r, 4, d, F_TX, bd=True); put(ws, r, 5, dv, F_TX, align=CTR, bd=True)
    put(ws, r, 6, note, F_TX, align=WRAP, bd=True)
put(ws, 13, 1, '缺陷加权权重（由项目2缺陷表原始数据反推核对：集成102/验收28/总计130全部吻合；非常见10/5/2/1口径）', F_BT)
for i, (name, sym, val) in enumerate([('致命缺陷权重', 'w_fatal', 8), ('严重缺陷权重', 'w_severe', 4),
                                      ('一般缺陷权重', 'w_normal', 1), ('细微缺陷权重', 'w_minor', 0.5)]):
    r = 14 + i
    put(ws, r, 1, name, F_TX, bd=True); put(ws, r, 2, sym, F_NT, bd=True)
    put(ws, r, 3, val, F_IN, fmt=N1, align=CTR, bd=True)
    put(ws, r, 6, '缺陷加权数=Σ(各级数量×权重)', F_TX, bd=True)
put(ws, 19, 1, '单位换算：1人天 =', F_TX, bd=True); put(ws, 19, 2, 'hours_per_day', F_NT, bd=True)
put(ws, 19, 3, 8, F_IN, fmt=INT, align=CTR, bd=True); put(ws, 19, 4, '人时', F_TX, bd=True)
put(ws, 19, 6, '功能点耗时率换算用；1人月=20人天', F_TX, bd=True)
put(ws, 21, 1, '红绿灯判定规则', F_BT)
c = put(ws, 22, 1, MON['statusRules'], F_TX, align=WRAP)
ws.merge_cells('A22:F22'); ws.row_dimensions[22].height = est_h([MON['statusRules']], [75])

PR = "'01参数'!"   # param sheet ref prefix

# ============================== 02 项目1度量 ==============================
ws = wb.create_sheet('02项目1度量')
widths(ws, {'A': 26, 'B': 14, 'C': 14, 'D': 14, 'E': 14, 'F': 30, 'G': 15, 'H': 12, 'I': 12})
ws.freeze_panes = 'A3'
put(ws, 1, 1, '项目1度量数据（度量表1）  周期：%s  实际规模256FP' % P1['period'], F_TI)
put(ws, 2, 1, '蓝字=原表计划/实际原值（可修改）；黑字=公式；总计行与偏差率一律公式计算。', F_NT)
# ---- 规模 ----
put(ws, 4, 1, '一、规模度量（各阶段规模对比，单位：功能点FP）', F_BT)
header_row(ws, 5, ['编号', '阶段', '估算规模(FP)', '实际规模(FP)', '规模偏差率', '说明'])
for i, ph in enumerate(P1['sizeFP']['byPhase']):
    r = 6 + i
    put(ws, r, 1, i + 1, F_TX, INT, CTR, bd=True)
    put(ws, r, 2, ph['phase'], F_TX, bd=True)
    put(ws, r, 3, ph['estimated'], F_IN, INT, CTR, bd=True)
    put(ws, r, 4, ph['actual'], F_IN, INT, CTR, bd=True)
    put(ws, r, 5, '=IF($C%d="","",($D%d-$C%d)/$C%d)' % (r, r, r, r), F_FM, PCT, CTR, bd=True)
    put(ws, r, 6, '' if ph['estimated'] is not None else '验收结项无估算值，偏差率空缺（估算187FP→实际256FP，跳变约27%）',
        F_NT, align=WRAP, bd=True)
put(ws, 12, 1, '结论（原表）：' + P1['conclusions']['size'], F_TX, align=WRAP)
ws.merge_cells('A12:F12'); ws.row_dimensions[12].height = est_h([P1['conclusions']['size']], [48])
# ---- 工作量 ----
put(ws, 14, 1, '二、工作量度量（单位：人天；1人月=20人天，1人天=8人时；偏差率以需求阶段估计值为基准）', F_BT)
header_row(ws, 15, ['编号', '工作量分类', '活动/阶段', '估计值-策划阶段', '估计值-需求阶段', '实际值', '工作量偏差率'])
for i, e in enumerate(P1['effort'][:12]):
    r = 16 + i
    put(ws, r, 1, e['no'], F_TX, INT, CTR, bd=True)
    put(ws, r, 2, e['category'], F_TX, bd=True)
    put(ws, r, 3, e['activity'], F_TX, bd=True)
    put(ws, r, 4, e['estPlan'], F_IN, N2, None, bd=True)
    put(ws, r, 5, e['estReq'], F_IN, N2, None, bd=True)
    put(ws, r, 6, e['actual'], F_IN, N2, None, bd=True)
    put(ws, r, 7, '=($F%d-$E%d)/$E%d' % (r, r, r), F_FM, PCT, CTR, bd=True)
put(ws, 28, 1, '所有活动阶段总计', F_B, align=CTR, fill=FILL_TOT, bd=True)
ws.merge_cells('A28:C28')
for col, letter in ((4, 'D'), (5, 'E'), (6, 'F')):
    put(ws, 28, col, '=SUM(%s16:%s27)' % (letter, letter), F_FMB, N2, None, FILL_TOT, bd=True)
put(ws, 28, 7, '=($F$28-$E$28)/$E$28', F_FMB, PCT, CTR, FILL_TOT, bd=True)
put(ws, 29, 1, '注：' + P1['effortNote'], F_NT)
put(ws, 30, 1, '结论（原表）：' + P1['conclusions']['effort'], F_TX, align=WRAP)
ws.merge_cells('A30:G30'); ws.row_dimensions[30].height = est_h([P1['conclusions']['effort']], [55])
# ---- 进度 ----
put(ws, 32, 1, '三、进度度量（单位：天）', F_BT)
header_row(ws, 33, ['阶段', '计划开始', '计划完成', '实际开始', '实际完成', '计划完成天数', '实际完成天数', '偏差天数', '进度偏差率'])
for i, s in enumerate(P1['schedule']):
    r = 34 + i
    put(ws, r, 1, s['phase'], F_TX, bd=True)
    put(ws, r, 2, s['planStart'], F_IN, align=CTR, bd=True)
    put(ws, r, 3, s['planEnd'], F_IN, align=CTR, bd=True)
    put(ws, r, 4, s['actStart'], F_IN, align=CTR, bd=True)
    put(ws, r, 5, s['actEnd'], F_IN, align=CTR, bd=True)
    put(ws, r, 6, s['planDays'], F_IN, INT, CTR, bd=True)
    put(ws, r, 7, s['actDays'], F_IN, INT, CTR, bd=True)
    put(ws, r, 8, '=$G%d-$F%d' % (r, r), F_FM, INT, CTR, bd=True)
    put(ws, r, 9, '=$H%d/$F%d' % (r, r), F_FM, PCT, CTR, bd=True)
put(ws, 40, 1, '总计（本工作簿添加）', F_B, align=CTR, fill=FILL_TOT, bd=True)
put(ws, 40, 6, '=SUM(F34:F39)', F_FMB, INT, CTR, FILL_TOT, bd=True)
put(ws, 40, 7, '=SUM(G34:G39)', F_FMB, INT, CTR, FILL_TOT, bd=True)
put(ws, 40, 8, '=$G$40-$F$40', F_FMB, INT, CTR, FILL_TOT, bd=True)
put(ws, 40, 9, '=$H$40/$F$40', F_FMB, PCT, CTR, FILL_TOT, bd=True)
put(ws, 41, 1, '备注：偏差天数=实际完成天数-计划完成天数；原表编码阶段实际完成(05-14)晚于计划(05-13)但天数均记61天、偏差记0，数据可信度见00说明。', F_NT)
put(ws, 42, 1, '结论（原表）：' + P1['conclusions']['schedule'], F_TX, align=WRAP)
ws.merge_cells('A42:I42'); ws.row_dimensions[42].height = est_h([P1['conclusions']['schedule']], [60])
# ---- 成本 ----
put(ws, 44, 1, '四、成本度量（单位：元；估计源《WBS估算表》，实际源《里程碑状态报告》）', F_BT)
header_row(ws, 45, ['里程碑', '估算成本(元)', '实际成本(元)', '成本偏差率'])
for i, cst in enumerate(P1['cost'][:6]):
    r = 46 + i
    put(ws, r, 1, cst['milestone'], F_TX, bd=True)
    put(ws, r, 2, cst['estimated'], F_IN, MONEY, None, bd=True)
    put(ws, r, 3, cst['actual'], F_IN, MONEY, None, bd=True)
    put(ws, r, 4, '=($C%d-$B%d)/$B%d' % (r, r, r), F_FM, PCT, CTR, bd=True)
put(ws, 52, 1, '总计', F_B, align=CTR, fill=FILL_TOT, bd=True)
put(ws, 52, 2, '=SUM(B46:B51)', F_FMB, MONEY, None, FILL_TOT, bd=True)
put(ws, 52, 3, '=SUM(C46:C51)', F_FMB, MONEY, None, FILL_TOT, bd=True)
put(ws, 52, 4, '=($C$52-$B$52)/$B$52', F_FMB, PCT, CTR, FILL_TOT, bd=True)
put(ws, 53, 1, '结论（原表）：' + P1['conclusions']['cost'], F_TX, align=WRAP)
ws.merge_cells('A53:D53'); ws.row_dimensions[53].height = est_h([P1['conclusions']['cost']], [42])
# ---- 关键派生指标 ----
put(ws, 55, 1, '五、关键派生指标（全部公式自动计算，供04汇总对比/05监控报告引用）', F_BT)
kpi1 = [
    ('实际规模（FP）', '=$D$11', INT, '取验收结项阶段实际规模'),
    ('总工作量-需求阶段估计（人天）', '=$E$28', N2, ''),
    ('总工作量-实际（人天）', '=$F$28', N2, ''),
    ('工作量偏差率', '=$G$28', PCT, '基准=需求阶段估计'),
    ('生产率（FP/人天）', '=$B$56/$B$58', N3, '＝实际规模/实际总工作量'),
    ('功能点耗时率（人时/FP）', '=$B$58*' + PR + '$C$19/$B$56', N2, '原表未给出，同口径推算值'),
    ('总工期-计划（天）', '=$F$40', INT, '阶段天数合计'),
    ('总工期-实际（天）', '=$G$40', INT, '日历跨度为127天，见00说明'),
    ('进度偏差率', '=$I$40', PCT, ''),
    ('总成本-估算（元）', '=$B$52', MONEY, ''),
    ('总成本-实际（元）', '=$C$52', MONEY, '疑按估算折算而非实采，见00说明'),
    ('成本偏差率', '=$D$52', PCT, ''),
    ('人天成本-实际（元/人天）', '=$B$66/$B$58', MONEY, '＝实际总成本/实际总工作量'),
    ('单功能点成本-实际（元/FP）', '=$B$66/$B$56', MONEY, '＝实际总成本/实际规模'),
]
for i, (name, fx, fmt, note) in enumerate(kpi1):
    r = 56 + i
    put(ws, r, 1, name, F_TX, bd=True)
    put(ws, r, 2, fx, F_FM, fmt, CTR, bd=True)
    put(ws, r, 3, note, F_NT, align=WRAP, bd=True)
    ws.merge_cells('C%d:F%d' % (r, r))

# ============================== 03 项目2度量 ==============================
ws = wb.create_sheet('03项目2度量')
widths(ws, {'A': 26, 'B': 13, 'C': 13, 'D': 13, 'E': 13, 'F': 12, 'G': 14, 'H': 12,
            'I': 12, 'J': 13, 'K': 12, 'L': 10, 'M': 32})
ws.freeze_panes = 'A3'
put(ws, 1, 1, '项目2度量数据（度量表2，编号KSSH-1102，V1.5受控）  周期：%s' % P2['period'], F_TI)
put(ws, 2, 1, '蓝字=原表计划/实际原值；黑字=公式（分布比例、累计偏差率、加权数、密度、探测率、修复率、效率均为公式）；绿字=引用01参数。', F_NT)
# ---- 工作量 ----
put(ws, 4, 1, '一、工作量度量（单位：人天；计划源项目估算表，实际源里程碑报告）', F_BT)
header_row(ws, 5, ['阶段', '计划工作量(人天)', '实际工作量(人天)', '分布比例', '累计工作量偏差率', '控制范围', '结果'])
for i, e in enumerate(P2['effort'][:6]):
    r = 6 + i
    put(ws, r, 1, e['phase'], F_TX, bd=True)
    put(ws, r, 2, e['plan'], F_IN, N2, None, bd=True)
    put(ws, r, 3, e['actual'], F_IN, N2, None, bd=True)
    put(ws, r, 4, '=$C%d/$C$12' % r, F_FM, PCT, CTR, bd=True)
    put(ws, r, 5, '=(SUM($C$6:$C%d)-SUM($B$6:$B%d))/SUM($B$6:$B%d)' % (r, r, r), F_FM, PCT, CTR, bd=True)
put(ws, 12, 1, '总计', F_B, align=CTR, fill=FILL_TOT, bd=True)
put(ws, 12, 2, '=SUM(B6:B11)', F_FMB, N2, None, FILL_TOT, bd=True)
put(ws, 12, 3, '=SUM(C6:C11)', F_FMB, N2, None, FILL_TOT, bd=True)
put(ws, 12, 4, '=SUM(D6:D11)', F_FMB, PCT, CTR, FILL_TOT, bd=True)
put(ws, 12, 5, '=($C$12-$B$12)/$B$12', F_FMB, PCT, CTR, FILL_TOT, bd=True)
put(ws, 6, 6, '="±"&TEXT(%s$C$4,"0%%")' % PR, F_XR, align=CTR, bd=True); ws.merge_cells('F6:F12')
put(ws, 6, 7, '结果（原表）：' + P2['effortResult'], F_TX, align=WRAP, bd=True); ws.merge_cells('G6:G12')
# ---- 规模及效率 ----
put(ws, 15, 1, '二、规模及效率', F_BT)
header_row(ws, 16, ['估计规模(FP)', '实际规模(FP)', '规模偏差率', '功能点耗时率(人时/FP)', '目标范围', '结果'])
put(ws, 17, 1, P2['sizeFP']['estimated'], F_IN, INT, CTR, bd=True)
put(ws, 17, 2, P2['sizeFP']['actual'], F_IN, INT, CTR, bd=True)
put(ws, 17, 3, '=($B$17-$A$17)/$A$17', F_FM, PCT, CTR, bd=True)
put(ws, 17, 4, '=$C$12*%s$C$19/$B$17' % PR, F_FM, N2, CTR, bd=True)
put(ws, 17, 5, '="≤"&TEXT(%s$C$9,"0.00")' % PR, F_XR, align=CTR, bd=True)
put(ws, 17, 6, '结果（原表）：' + P2['conclusions']['size'], F_TX, align=WRAP, bd=True)
ws.row_dimensions[17].height = 30
# ---- 工期 ----
put(ws, 19, 1, '三、工期度量（单位：天，工作日口径已扣节假日；日历跨度计划189天/实际202天）', F_BT)
header_row(ws, 20, ['阶段', '计划开始', '计划完成', '实际开始', '实际完成', '计划工期(天)', '计划累计(天)',
                    '实际工期(天)', '实际累计(天)', '累计工期偏差率', '控制范围', '结果'])
for i, s in enumerate(P2['schedule'][:6]):
    r = 21 + i
    put(ws, r, 1, s['phase'], F_TX, bd=True)
    put(ws, r, 2, s['planStart'], F_IN, align=CTR, bd=True)
    put(ws, r, 3, s['planEnd'], F_IN, align=CTR, bd=True)
    put(ws, r, 4, s['actStart'], F_IN, align=CTR, bd=True)
    put(ws, r, 5, s['actEnd'], F_IN, align=CTR, bd=True)
    put(ws, r, 6, s['planDays'], F_IN, INT, CTR, bd=True)
    put(ws, r, 7, '=SUM($F$21:$F%d)' % r, F_FM, INT, CTR, bd=True)
    put(ws, r, 8, s['actDays'], F_IN, INT, CTR, bd=True)
    put(ws, r, 9, '=SUM($H$21:$H%d)' % r, F_FM, INT, CTR, bd=True)
    put(ws, r, 10, '=($I%d-$G%d)/$G%d' % (r, r, r), F_FM, PCT, CTR, bd=True)
put(ws, 27, 1, '总计', F_B, align=CTR, fill=FILL_TOT, bd=True)
put(ws, 27, 6, '=SUM(F21:F26)', F_FMB, INT, CTR, FILL_TOT, bd=True)
put(ws, 27, 7, '=$G$26', F_FMB, INT, CTR, FILL_TOT, bd=True)
put(ws, 27, 8, '=SUM(H21:H26)', F_FMB, INT, CTR, FILL_TOT, bd=True)
put(ws, 27, 9, '=$I$26', F_FMB, INT, CTR, FILL_TOT, bd=True)
put(ws, 27, 10, '=($I$27-$G$27)/$G$27', F_FMB, PCT, CTR, FILL_TOT, bd=True)
put(ws, 21, 11, '="±"&TEXT(%s$C$5,"0%%")' % PR, F_XR, align=CTR, bd=True); ws.merge_cells('K21:K27')
put(ws, 21, 12, '结果（原表）：' + P2['scheduleResult'], F_TX, align=WRAP, bd=True); ws.merge_cells('L21:L27')
put(ws, 28, 1, '注：' + P2['scheduleNote'], F_NT)
# ---- 成本 ----
put(ws, 30, 1, '四、成本度量（单位：元；计划成本=计划工作量×1200元/人天推导值，实际为独立采集）', F_BT)
header_row(ws, 31, ['阶段', '计划成本(元)', '实际成本(元)', '累计偏差率', '控制范围', '结果'])
for i, cst in enumerate(P2['cost'][:6]):
    r = 32 + i
    put(ws, r, 1, cst['phase'], F_TX, bd=True)
    put(ws, r, 2, cst['plan'], F_IN, MONEY, None, bd=True)
    put(ws, r, 3, cst['actual'], F_IN, MONEY, None, bd=True)
    put(ws, r, 4, '=(SUM($C$32:$C%d)-SUM($B$32:$B%d))/SUM($B$32:$B%d)' % (r, r, r), F_FM, PCT, CTR, bd=True)
put(ws, 38, 1, '总计', F_B, align=CTR, fill=FILL_TOT, bd=True)
put(ws, 38, 2, '=SUM(B32:B37)', F_FMB, MONEY, None, FILL_TOT, bd=True)
put(ws, 38, 3, '=SUM(C32:C37)', F_FMB, MONEY, None, FILL_TOT, bd=True)
put(ws, 38, 4, '=($C$38-$B$38)/$B$38', F_FMB, PCT, CTR, FILL_TOT, bd=True)
put(ws, 32, 5, '="±"&TEXT(%s$C$6,"0%%")' % PR, F_XR, align=CTR, bd=True); ws.merge_cells('E32:E38')
put(ws, 32, 6, '结果（原表）：' + P2['costResult'], F_TX, align=WRAP, bd=True); ws.merge_cells('F32:F38')
put(ws, 39, 1, '注：原表成本「结果」栏文字描述对象疑为工期（复制粘贴笔误），照录备考。', F_NT)
# ---- 缺陷 ----
put(ws, 41, 1, '五、测试缺陷度量（加权权重引用01参数：致命×8、严重×4、一般×1、细微×0.5，由原始数据反推核对一致）', F_BT)
header_row(ws, 42, ['测试类型', '致命', '严重', '一般', '细微', '检出总数(个)', '关闭缺陷数(个)', '缺陷修复率',
                    '缺陷加权数', '缺陷密度(个/FP)', '缺陷探测率', '目标值', '结果'])
for i, t in enumerate(P2['defects']['byTest'][:2]):
    r = 43 + i
    put(ws, r, 1, t['type'], F_TX, bd=True)
    for j, k in enumerate(['fatal', 'severe', 'normal', 'minor']):
        put(ws, r, 2 + j, t[k], F_IN, INT, CTR, bd=True)
    put(ws, r, 6, '=SUM($B%d:$E%d)' % (r, r), F_FM, INT, CTR, bd=True)
    put(ws, r, 7, t['closed'], F_IN, INT, CTR, bd=True)
    put(ws, r, 8, '=$G%d/$F%d' % (r, r), F_FM, PCT, CTR, bd=True)
    put(ws, r, 9, '=$B{r}*{p}$C$14+$C{r}*{p}$C$15+$D{r}*{p}$C$16+$E{r}*{p}$C$17'.format(r=r, p=PR), F_FM, N1, CTR, bd=True)
    put(ws, r, 10, '=$I%d/$B$17' % r, F_FM, N3, CTR, bd=True)
    if i == 0:
        put(ws, r, 11, '=$I$43/$I$45', F_FM, PCT, CTR, bd=True)
    else:
        put(ws, r, 11, '', F_TX, align=CTR, bd=True)
put(ws, 45, 1, '总计', F_B, align=CTR, fill=FILL_TOT, bd=True)
for j, letter in enumerate(['B', 'C', 'D', 'E']):
    put(ws, 45, 2 + j, '=SUM(%s43:%s44)' % (letter, letter), F_FMB, INT, CTR, FILL_TOT, bd=True)
put(ws, 45, 6, '=SUM($B$45:$E$45)', F_FMB, INT, CTR, FILL_TOT, bd=True)
put(ws, 45, 7, '=SUM(G43:G44)', F_FMB, INT, CTR, FILL_TOT, bd=True)
put(ws, 45, 8, '=$G$45/$F$45', F_FMB, PCT, CTR, FILL_TOT, bd=True)
put(ws, 45, 9, '=$B$45*{p}$C$14+$C$45*{p}$C$15+$D$45*{p}$C$16+$E$45*{p}$C$17'.format(p=PR), F_FMB, N1, CTR, FILL_TOT, bd=True)
put(ws, 45, 10, '=$I$45/$B$17', F_FMB, N3, CTR, FILL_TOT, bd=True)
put(ws, 45, 11, '', F_TX, fill=FILL_TOT, bd=True)
put(ws, 43, 12, '="≥"&TEXT(%s$C$8,"0%%")' % PR, F_XR, align=CTR, bd=True); ws.merge_cells('L43:L45')
put(ws, 43, 13, '结果（原表）：' + P2['defects']['detectRateResult'], F_TX, align=WRAP, bd=True); ws.merge_cells('M43:M45')
put(ws, 46, 1, '注：缺陷探测率=集成与系统测试加权数/总加权数（仅该行计算，验收测试与总计行原表空缺）；缺陷密度=加权数/实际规模FP。', F_NT)
# ---- 评审缺陷 ----
put(ws, 48, 1, '六、评审缺陷度量', F_BT)
header_row(ws, 49, ['严重', '一般', '轻微', '缺陷总数', '已关闭', '评审用时(人时)', '评审效率(个/人时)', '目标范围', '结果'])
rv = P2['defects']['reviewDefects']
put(ws, 50, 1, rv['severe'], F_IN, INT, CTR, bd=True)
put(ws, 50, 2, rv['normal'], F_IN, INT, CTR, bd=True)
put(ws, 50, 3, rv['minor'], F_IN, INT, CTR, bd=True)
put(ws, 50, 4, '=SUM($A$50:$C$50)', F_FM, INT, CTR, bd=True)
put(ws, 50, 5, rv['closed'], F_IN, INT, CTR, bd=True)
put(ws, 50, 6, rv['reviewHours'], F_IN, INT, CTR, bd=True)
put(ws, 50, 7, '=$D$50/$F$50', F_FM, N3, CTR, bd=True)
put(ws, 50, 8, '="≥"&TEXT(%s$C$10,"0.00")' % PR, F_XR, align=CTR, bd=True)
put(ws, 50, 9, '结果（原表）：' + rv['result'], F_TX, align=WRAP, bd=True)
ws.row_dimensions[50].height = 30
# ---- 关键派生指标 ----
put(ws, 52, 1, '七、关键派生指标（全部公式自动计算，供04汇总对比/05监控报告引用）', F_BT)
kpi2 = [
    ('实际规模（FP）', '=$B$17', INT, ''),
    ('总工作量-计划（人天）', '=$B$12', N2, ''),
    ('总工作量-实际（人天）', '=$C$12', N2, ''),
    ('工作量偏差率', '=$E$12', PCT, ''),
    ('规模偏差率', '=$C$17', PCT, ''),
    ('生产率（FP/人天）', '=$B$53/$B$55', N3, '＝实际规模/实际总工作量'),
    ('功能点耗时率（人时/FP）', '=$D$17', N2, ''),
    ('总工期-计划（天）', '=$F$27', INT, '工作日口径'),
    ('总工期-实际（天）', '=$H$27', INT, '工作日口径'),
    ('工期偏差率（累计）', '=$J$27', PCT, '总计行取交付验收阶段累计值'),
    ('总成本-计划（元）', '=$B$38', MONEY, '计划工作量×1200元/人天'),
    ('总成本-实际（元）', '=$C$38', MONEY, ''),
    ('成本偏差率（累计）', '=$D$38', PCT, ''),
    ('人天成本-实际（元/人天）', '=$B$64/$B$55', MONEY, ''),
    ('单功能点成本-实际（元/FP）', '=$B$64/$B$53', MONEY, ''),
    ('缺陷密度（个/FP）', '=$J$45', N3, '＝总加权数/实际规模'),
    ('缺陷探测率', '=$K$43', PCT, '＝集成系统测试加权/总加权'),
    ('评审效率（个/人时）', '=$G$50', N3, ''),
]
for i, (name, fx, fmt, note) in enumerate(kpi2):
    r = 53 + i
    put(ws, r, 1, name, F_TX, bd=True)
    put(ws, r, 2, fx, F_FM, fmt, CTR, bd=True)
    put(ws, r, 3, note, F_NT, align=WRAP, bd=True)
    ws.merge_cells('C%d:E%d' % (r, r))

# ============================== 04 汇总对比 ==============================
ws = wb.create_sheet('04汇总对比')
widths(ws, {'A': 26, 'B': 12, 'C': 16, 'D': 16, 'E': 58})
ws.freeze_panes = 'A5'
put(ws, 1, 1, '跨项目度量汇总对比', F_TI)
put(ws, 2, 1, '项目1/项目2数值全部为跨表引用公式（绿字），随02/03表输入联动；“—”表示原表无该度量。', F_NT)
header_row(ws, 4, ['指标', '单位', '项目1（256FP）', '项目2（327FP）', '差异说明'])
Q2 = "'02项目1度量'!"; Q3 = "'03项目2度量'!"
cmp_rows = [
    ('项目规模（实际）', 'FP', '=%s$B$56' % Q2, '=%s$B$53' % Q3, INT),
    ('总工作量（实际）', '人天', '=%s$B$58' % Q2, '=%s$B$55' % Q3, N2),
    ('总工期（实际）', '天', '=%s$B$63' % Q2, '=%s$B$61' % Q3, INT),
    ('总成本（估算/计划）', '元', '=%s$B$65' % Q2, '=%s$B$63' % Q3, MONEY),
    ('总成本（实际）', '元', '=%s$B$66' % Q2, '=%s$B$64' % Q3, MONEY),
    ('生产率', 'FP/人天', '=%s$B$60' % Q2, '=%s$B$58' % Q3, N3),
    ('功能点耗时率', '人时/FP', '=%s$B$61' % Q2, '=%s$B$59' % Q3, N2),
    ('人天成本（实际）', '元/人天', '=%s$B$68' % Q2, '=%s$B$66' % Q3, MONEY),
    ('单功能点成本（实际）', '元/FP', '=%s$B$69' % Q2, '=%s$B$67' % Q3, MONEY),
    ('工作量偏差率', '%', '=%s$B$59' % Q2, '=%s$B$56' % Q3, PCT),
    ('进度/工期偏差率', '%', '=%s$B$64' % Q2, '=%s$B$62' % Q3, PCT),
    ('成本偏差率', '%', '=%s$B$67' % Q2, '=%s$B$65' % Q3, PCT),
    ('规模偏差率', '%', '—', '=%s$B$57' % Q3, PCT),
    ('缺陷密度', '个/FP', '—', '=%s$B$68' % Q3, N3),
    ('缺陷探测率', '%', '—', '=%s$B$69' % Q3, PCT),
]
notes4 = [c['note'] for c in MET['crossProject']['comparison']]
for i, (name, unit, f1, f2, fmt) in enumerate(cmp_rows):
    r = 5 + i
    put(ws, r, 1, name, F_TX, bd=True)
    put(ws, r, 2, unit, F_TX, align=CTR, bd=True)
    put(ws, r, 3, f1, F_XR if str(f1).startswith('=') else F_NT, fmt if str(f1).startswith('=') else None, CTR, bd=True)
    put(ws, r, 4, f2, F_XR, fmt, CTR, bd=True)
    put(ws, r, 5, notes4[i], F_TX, align=WRAP, bd=True)
    ws.row_dimensions[r].height = est_h([notes4[i]], [28])

# ============================== 05 监控报告 ==============================
ws = wb.create_sheet('05监控报告')
widths(ws, {'A': 5, 'B': 7, 'C': 20, 'D': 11, 'E': 11, 'F': 10, 'G': 11, 'H': 62, 'I': 20, 'J': 20})
ws.freeze_panes = 'A6'
put(ws, 1, 1, '项目度量红绿灯监控报告', F_TI)
put(ws, 2, 1, MON['statusRules'], F_TX, align=WRAP); ws.merge_cells('B2:H2' if False else 'A2:H2')
ws.row_dimensions[2].height = est_h([MON['statusRules']], [80])
put(ws, 3, 1, '指标值与阈值均为跨表引用（绿字）；状态列为公式，随01参数阈值联动；最右两列为自动结论的辅助记录列。', F_NT)
put(ws, 4, 1, '判定公式：偏差型 =IF(ABS(值)<=阈值×系数,"● 正常",IF(ABS(值)<=阈值,"▲ 预警","✖ 超限"))；达标型按≥/≤方向判达标。', F_NT)
header_row(ws, 5, ['序号', '项目', '维度/指标', '指标值', '阈值/目标', '判定类型', '状态', '结论建议', '超限记录', '预警记录'])
mon_rows = [
    ('P1', '工作量偏差率', '=%s$B$59' % Q2, PR + '$C$4', PCT, 'dev',
     '实际低于估算4.6%，绿区；注意P1偏差基准为需求阶段估计值'),
    ('P1', '进度偏差率', '=%s$B$64' % Q2, PR + '$C$5', PCT, 'dev',
     '全程记0偏差，绿区；但编码阶段实际完成晚于计划1天仍记0，疑被抹平，可信度存疑'),
    ('P1', '成本偏差率', '=%s$B$67' % Q2, PR + '$C$6', PCT, 'dev',
     '节支4.4%，绿区；各里程碑偏差率高度一致，实际成本疑按估算折算，建议实采核对'),
    ('P1', '功能点耗时率', '=%s$B$61' % Q2, PR + '$C$9', N2, 'le',
     '7.36人时/FP为推算值（原表未给出），略超7.15目标；横向对比需先统一口径'),
    ('P2', '工作量偏差率', '=%s$B$56' % Q3, PR + '$C$4', PCT, 'dev',
     '-1.6%，绿区，估算准确，各阶段分布比例合理'),
    ('P2', '规模偏差率', '=%s$B$57' % Q3, PR + '$C$7', PCT, 'dev',
     '-0.3%，绿区，规模基线控制好'),
    ('P2', '工期偏差率', '=%s$B$62' % Q3, PR + '$C$5', PCT, 'dev',
     '累计+7.4%，绿区；系统设计阶段曾达10%触及黄区后收敛，设计/编码阶段需设检查点'),
    ('P2', '成本偏差率', '=%s$B$65' % Q3, PR + '$C$6', PCT, 'dev',
     '超支+4.9%，绿区；实际成本全程高于计划（最高+6%），更大规模项目可能放大为超限'),
    ('P2', '功能点耗时率', '=%s$B$59' % Q3, PR + '$C$9', N2, 'le',
     '7.07≤7.15，达标'),
    ('P2', '缺陷探测率', '=%s$B$69' % Q3, PR + '$C$8', PCT, 'ge',
     '78.5%≥75%，达标；缺陷加权集中于集成与系统测试（致命2/严重15），评审与单测宜前移'),
    ('P2', '评审效率', '=%s$B$70' % Q3, PR + '$C$10', N3, 'ge',
     '0.758≥0.75，达标'),
]
for i, (prj, dim, vref, tref, fmt, typ, concl) in enumerate(mon_rows):
    r = 6 + i
    put(ws, r, 1, i + 1, F_TX, INT, CTR, bd=True)
    put(ws, r, 2, prj, F_TX, align=CTR, bd=True)
    put(ws, r, 3, dim, F_TX, bd=True)
    put(ws, r, 4, vref, F_XR, fmt, CTR, bd=True)
    put(ws, r, 5, '=' + tref, F_XR, fmt, CTR, bd=True)
    if typ == 'dev':
        put(ws, r, 6, '偏差型±', F_TX, align=CTR, bd=True)
        fx = '=IF(ABS($D{r})<=$E{r}*{p}$C$11,"● 正常",IF(ABS($D{r})<=$E{r},"▲ 预警","✖ 超限"))'.format(r=r, p=PR)
    elif typ == 'ge':
        put(ws, r, 6, '达标型≥', F_TX, align=CTR, bd=True)
        fx = '=IF($D{r}>=$E{r},"● 达标","✖ 未达标")'.format(r=r)
    else:
        put(ws, r, 6, '达标型≤', F_TX, align=CTR, bd=True)
        fx = '=IF($D{r}<=$E{r},"● 达标","✖ 未达标")'.format(r=r)
    put(ws, r, 7, fx, F_FMB, None, CTR, bd=True)
    put(ws, r, 8, concl, F_TX, align=WRAP, bd=True)
    put(ws, r, 9, '=IF(LEFT($G%d,1)="✖",$B%d&$C%d,"")' % (r, r, r), F_NT, align=CTR, bd=True)
    put(ws, r, 10, '=IF(LEFT($G%d,1)="▲",$B%d&$C%d,"")' % (r, r, r), F_NT, align=CTR, bd=True)
    ws.row_dimensions[r].height = est_h([concl], [28], min_h=26)
# 条件格式（状态列）
rng = 'G6:G16'
ws.conditional_formatting.add(rng, FormulaRule(
    formula=['NOT(ISERROR(SEARCH("●",$G6)))'], font=Font(name=AR, size=SZ, bold=True, color='0CA30C')))
ws.conditional_formatting.add(rng, FormulaRule(
    formula=['NOT(ISERROR(SEARCH("▲",$G6)))'], font=Font(name=AR, size=SZ, bold=True, color='B45309'),
    fill=PatternFill('solid', start_color='FFF6E5')))
ws.conditional_formatting.add(rng, FormulaRule(
    formula=['NOT(ISERROR(SEARCH("✖",$G6)))'], font=Font(name=AR, size=SZ, bold=True, color='D03B3B'),
    fill=PatternFill('solid', start_color='FDEAEA')))
# 统计与自动结论
put(ws, 18, 1, '自动统计', F_BT)
put(ws, 18, 2, '● 正常/达标', F_TX, align=CTR); ws.merge_cells('B18:C18')
put(ws, 18, 4, '=SUMPRODUCT(--(LEFT($G$6:$G$16,1)="●"))', F_FMB, INT, CTR)
put(ws, 18, 5, '▲ 预警', F_TX, align=CTR)
put(ws, 18, 6, '=SUMPRODUCT(--(LEFT($G$6:$G$16,1)="▲"))', F_FMB, INT, CTR)
put(ws, 18, 7, '✖ 超限/未达标', F_TX, align=CTR)
put(ws, 18, 8, '=SUMPRODUCT(--(LEFT($G$6:$G$16,1)="✖"))', F_FMB, INT, None)
concl_fx = ('="本次监控共 "&COUNTA($G$6:$G$16)&" 项指标：正常/达标 "&$D$18&" 项、预警 "&$F$18&" 项、超限/未达标 "&$H$18&" 项。"'
            '&IF($H$18=0,"无超限/未达标项。","超限/未达标项："&_xlfn.TEXTJOIN("、",TRUE,$I$6:$I$16)&"。")'
            '&IF($F$18=0,"","预警项："&_xlfn.TEXTJOIN("、",TRUE,$J$6:$J$16)&"。")')
put(ws, 19, 1, '自动结论', F_BT)
put(ws, 19, 2, concl_fx, F_FMB, None, WRAP); ws.merge_cells('B19:H19'); ws.row_dimensions[19].height = 34
put(ws, 21, 1, '总体研判（来自分析报告）', F_BT)
put(ws, 22, 2, MON['summary'], F_TX, align=WRAP); ws.merge_cells('B22:H22')
ws.row_dimensions[22].height = est_h([MON['summary']], [80])
put(ws, 24, 1, '主要风险', F_BT)
rr = 25
for i, t in enumerate(MON['risks'], 1):
    put(ws, rr, 1, 'R%d' % i, F_TX, align=CTR)
    put(ws, rr, 2, t, F_TX, align=WRAP); ws.merge_cells('B%d:H%d' % (rr, rr))
    ws.row_dimensions[rr].height = est_h([t], [80]); rr += 1
rr += 1
put(ws, rr, 1, '改进建议', F_BT); rr += 1
for i, t in enumerate(MON['suggestions'], 1):
    put(ws, rr, 1, 'S%d' % i, F_TX, align=CTR)
    put(ws, rr, 2, t, F_TX, align=WRAP); ws.merge_cells('B%d:H%d' % (rr, rr))
    ws.row_dimensions[rr].height = est_h([t], [80]); rr += 1

# ============================== 06 功能分解 ==============================
ws = wb.create_sheet('06功能分解')
widths(ws, {'A': 9, 'B': 14, 'C': 9, 'D': 15, 'E': 7, 'F': 13, 'G': 9, 'H': 70, 'I': 38, 'J': 38, 'K': 48, 'L': 38})
put(ws, 1, 1, '《供水营销管理系统》功能分解（WBS）全清单　来源：需求规格说明书（srs_full.txt）', F_TI)
pmapc = {1: 'P1', 2: 'P2', 5: 'P3'}
pmapl = {1: '高（核心必备）', 2: '中高', 5: '一般'}
put(ws, 3, 1, '模块数', F_B); put(ws, 3, 2, '=SUMPRODUCT(1/COUNTIF($A$9:$A$35,$A$9:$A$35))', F_FMB, INT, CTR)
put(ws, 3, 3, '功能数', F_B); put(ws, 3, 4, '=COUNTA($C$9:$C$35)', F_FMB, INT, CTR)
put(ws, 3, 5, 'P1数', F_B); put(ws, 3, 6, '=COUNTIF($E$9:$E$35,"P1")', F_FMB, INT, CTR)
put(ws, 4, 5, 'P2数', F_B); put(ws, 4, 6, '=COUNTIF($E$9:$E$35,"P2")', F_FMB, INT, CTR)
put(ws, 5, 5, 'P3数', F_B); put(ws, 5, 6, '=COUNTIF($E$9:$E$35,"P3")', F_FMB, INT, CTR)
put(ws, 4, 1, '优先级映射：SRS优先级1→P1 高（核心必备）；2→P2 中高；5→P3 一般', F_NT); ws.merge_cells('A4:D4')
put(ws, 5, 1, '统计为COUNTA/COUNTIF公式，随清单增删联动', F_NT); ws.merge_cells('A5:D5')
put(ws, 7, 1, '（首行已冻结，表头行已开启自动筛选）', F_NT)
header_row(ws, 8, ['模块编号', '模块名称', '功能编号', '功能名称', '优先级', '优先级说明', 'SRS章节',
                   '细化功能描述', '主要输入', '主要输出', '验收要点', '备注问题'])
r = 9
for m in DEC['modules']:
    for fn in m['functions']:
        pr_ = fn.get('priority')
        notes = fn.get('notes')
        if isinstance(notes, list):
            notes = '\n'.join(str(x) for x in notes)
        ins = '；'.join(fn.get('inputs', []) or [])
        outs = '；'.join(fn.get('outputs', []) or [])
        acc = '\n'.join(fn.get('acceptance', []) or [])
        put(ws, r, 1, m['id'], F_TX, align=CTR, bd=True)
        put(ws, r, 2, m['name'], F_TX, align=WRAP, bd=True)
        put(ws, r, 3, fn['id'], F_TX, align=CTR, bd=True)
        put(ws, r, 4, fn['name'], F_TX, align=WRAP, bd=True)
        put(ws, r, 5, pmapc.get(pr_, ''), F_TX, align=CTR, bd=True)
        put(ws, r, 6, pmapl.get(pr_, ''), F_TX, align=CTR, bd=True)
        put(ws, r, 7, fn.get('srsRef', ''), F_TX, align=CTR, bd=True)
        put(ws, r, 8, fn.get('refinedDesc', ''), F_TX, align=WRAP, bd=True)
        put(ws, r, 9, ins, F_TX, align=WRAP, bd=True)
        put(ws, r, 10, outs, F_TX, align=WRAP, bd=True)
        put(ws, r, 11, acc, F_TX, align=WRAP, bd=True)
        put(ws, r, 12, notes or '', F_TX, align=WRAP, bd=True)
        ws.row_dimensions[r].height = est_h(
            [fn.get('refinedDesc', ''), ins, outs, acc, notes or ''], [34, 18, 18, 23, 18], min_h=28, max_h=300)
        r += 1
last = r - 1
ws.auto_filter.ref = 'A8:L%d' % last
ws.freeze_panes = 'A9'
r += 2
put(ws, r, 1, '范围缺口：SRS 3.1/3.2 提及但第4章未详述的模块（扩展建议，不计入上方27项清单）', F_BT); r += 1
header_row(ws, r, ['模块', 'SRS线索/缺失原因', '建议补充功能'], start=1); hdr = r; r += 1
for e in DEC['expansion']:
    fns = '\n'.join('· ' + x for x in e['suggestedFunctions'])
    put(ws, r, 1, e['module'], F_TX, align=WRAP, bd=True)
    put(ws, r, 2, e['reason'], F_TX, align=WRAP, bd=True)
    put(ws, r, 3, fns, F_TX, align=WRAP, bd=True); ws.merge_cells('C%d:H%d' % (r, r))
    ws.row_dimensions[r].height = est_h([e['reason'], fns], [7, 60], min_h=40, max_h=300)
    r += 1

# ============================== 07 质量检查 ==============================
ws = wb.create_sheet('07质量检查')
widths(ws, {'A': 8, 'B': 30, 'C': 13, 'D': 8, 'E': 44, 'F': 44, 'G': 44, 'H': 9})
put(ws, 1, 1, '《供水营销管理系统》SRS质量检查问题清单', F_TI)
put(ws, 2, 1, '总评：' + QUA['summary'], F_TX, align=WRAP); ws.merge_cells('A2:H2')
ws.row_dimensions[2].height = est_h([QUA['summary']], [95])
put(ws, 4, 1, '问题总数', F_B); put(ws, 4, 2, '=COUNTA($A$10:$A$50)', F_FMB, INT, None)
put(ws, 4, 3, '高', F_B, align=CTR); put(ws, 4, 4, '=COUNTIF($D$10:$D$50,"高")', F_FMB, INT, CTR)
put(ws, 4, 5, '中', F_B, align=CTR); put(ws, 4, 6, '=COUNTIF($D$10:$D$50,"中")', F_FMB, INT, CTR)
put(ws, 4, 7, '低', F_B, align=CTR); put(ws, 4, 8, '=COUNTIF($D$10:$D$50,"低")', F_FMB, INT, CTR)
cats = list(QUA['stats']['byCategory'].keys())
for i, cat in enumerate(cats):
    rr = 5 + i // 4
    cc = (i % 4) * 2 + 1
    put(ws, rr, cc, cat, F_TX, align=CTR)
    put(ws, rr, cc + 1, '=COUNTIF($C$10:$C$50,"%s")' % cat, F_FM, INT, CTR)
put(ws, 8, 1, '（统计均为COUNTIF公式；严重度列条件格式：高=红底白字、中=黄底、低=灰字；首行冻结+自动筛选）', F_NT)
header_row(ws, 9, ['编号', '章节位置', '类别', '严重度', '原文摘录', '问题说明', '修改建议', '状态'])
r = 10
for q in QUA['findings']:
    put(ws, r, 1, q['id'], F_TX, align=CTR, bd=True)
    put(ws, r, 2, q['location'], F_TX, align=WRAP, bd=True)
    put(ws, r, 3, q['category'], F_TX, align=CTR, bd=True)
    put(ws, r, 4, q['severity'], F_B, align=CTR, bd=True)
    put(ws, r, 5, q['quote'], F_TX, align=WRAP, bd=True)
    put(ws, r, 6, q['problem'], F_TX, align=WRAP, bd=True)
    put(ws, r, 7, q['suggestion'], F_TX, align=WRAP, bd=True)
    put(ws, r, 8, q['status'], F_TX, align=CTR, bd=True)
    ws.row_dimensions[r].height = est_h([q['quote'], q['problem'], q['suggestion'], q['location']],
                                        [21, 21, 21, 14], min_h=28, max_h=280)
    r += 1
last = r - 1
ws.auto_filter.ref = 'A9:H%d' % last
ws.freeze_panes = 'A10'
sev_rng = 'D10:D%d' % last
ws.conditional_formatting.add(sev_rng, CellIsRule(
    operator='equal', formula=['"高"'], fill=PatternFill('solid', start_color='D03B3B'),
    font=Font(name=AR, size=SZ, bold=True, color='FFFFFF')))
ws.conditional_formatting.add(sev_rng, CellIsRule(
    operator='equal', formula=['"中"'], fill=PatternFill('solid', start_color='FFE699'),
    font=Font(name=AR, size=SZ, bold=True, color='7F6000')))
ws.conditional_formatting.add(sev_rng, CellIsRule(
    operator='equal', formula=['"低"'], font=Font(name=AR, size=SZ, color='808080')))

wb.save(OUT)
print('saved:', OUT)
